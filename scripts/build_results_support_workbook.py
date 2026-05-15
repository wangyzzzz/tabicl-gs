from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs" / "results_support_workbook"
WORKBOOK_PATH = OUTPUT_DIR / "results_support_tables_20260512.xlsx"

MAIN_RESULTS_PATH = OUTPUT_DIR / "main_results_non_pig_fixed.csv"
COMPARE_ALL_PATH = OUTPUT_DIR / "compare_all_41_traits.csv"
MARKER_COUNT_PATH = OUTPUT_DIR / "marker_count_main_results.csv"
TABPFN_PATH = ROOT / "outputs" / "5.4-tabpfn-10k-8traits-fusion" / "tabpfn_10k_8traits_compare.csv"
ANALYSIS_TABLE_PATH = ROOT / "outputs" / "figures_20260506_results" / "analysis_table.csv"
SAMPLE_SIZE_DIR = OUTPUT_DIR / "sample_size_server_dir"
TABPFN_SUPPORT_DIR = OUTPUT_DIR / "tabpfn_fusion_server_dir"

SAMPLE_SIZE_TRAITS_8 = {
    ("cotton1245", "cotton_fibelo_17_18_cotton_fibelo_17_18"),
    ("cotton1245", "cotton_fiblen_17_18_cotton_fiblen_17_18"),
    ("rice529", "grain_weight"),
    ("rice529", "grain_width"),
    ("soybean951", "bbd_beijing_2013_bbd_beijing_2013"),
    ("soybean951", "lw_beijing_2013_lw_beijing_2013"),
    ("wheat406", "sl_e1"),
    ("wheat406", "sl_e2"),
}

R1_CASE_TRAITS = [
    ("rice529", "grain_weight"),
    ("rice529", "plant_height"),
    ("wheat406", "pl_e1"),
]

R2_SINGLE_CASE_TRAITS = [
    ("wheat406", "sl_e1"),
    ("rice529", "grain_weight"),
    ("rice529", "heading_date"),
    ("wheat406", "pl_e1"),
]


def require_path(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Missing required file: {path}")


def pct_diff(new_value: pd.Series, base_value: pd.Series) -> pd.Series:
    return (new_value - base_value) / base_value * 100.0


def round_frame(df: pd.DataFrame, digits: int) -> pd.DataFrame:
    out = df.copy()
    float_cols = out.select_dtypes(include=["float", "float64", "float32"]).columns
    if len(float_cols) > 0:
        out[float_cols] = out[float_cols].round(digits)
    return out


def best_of_three(df: pd.DataFrame, cols: Iterable[str]) -> tuple[pd.Series, pd.Series]:
    cols = list(cols)
    best_name = df[cols].idxmax(axis=1)
    best_value = df[cols].max(axis=1)
    return best_name, best_value


def load_sample_size_raw() -> pd.DataFrame:
    records: list[dict] = []
    for csv_path in sorted(SAMPLE_SIZE_DIR.glob("*/*/p_*/repeat_*/compare/compare_main.csv")):
        parts = csv_path.parts
        idx = parts.index("sample_size_server_dir")
        dataset = parts[idx + 1]
        trait_slug = parts[idx + 2]
        sample_fraction = parts[idx + 3].replace("p_", "")
        repeat = int(parts[idx + 4].replace("repeat_", ""))
        row = pd.read_csv(csv_path).iloc[0].to_dict()
        row["dataset"] = dataset
        row["trait_slug"] = trait_slug
        row["sample_fraction"] = sample_fraction
        row["sample_label"] = f"{int(float(sample_fraction) * 100)}%"
        row["repeat"] = repeat
        records.append(row)
    if not records:
        raise FileNotFoundError(f"No sample-size compare files found under {SAMPLE_SIZE_DIR}")
    return pd.DataFrame(records)


def build_readme_sheet() -> pd.DataFrame:
    rows = [
        ("context", "generated_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("context", "main_line", "5.4-duli-liudang"),
        ("context", "default_exclude", "pig3534"),
        ("context", "server_source", "server@GPU1:/home/server/code/git/TabICLv2-test"),
        ("context", "server_python", "/data/yes/envs/TabICLv2-GS/bin/python"),
        ("source", "main_results_non_pig_fixed", str(MAIN_RESULTS_PATH)),
        ("source", "compare_all_41_traits", str(COMPARE_ALL_PATH)),
        ("source", "analysis_table", str(ANALYSIS_TABLE_PATH)),
        ("source", "marker_count_main_results", str(MARKER_COUNT_PATH)),
        ("source", "tabpfn_10k_8traits_compare", str(TABPFN_PATH)),
        ("source", "sample_size_server_dir", str(SAMPLE_SIZE_DIR)),
        ("source", "tabpfn_fusion_server_dir", str(TABPFN_SUPPORT_DIR)),
        ("scope", "Result1-3", "36 non-pig traits"),
        ("scope", "Result4_sample_size", "8 traits x 3 sample fractions (20%, 60%, 100%)"),
        ("scope", "Result5_marker_count", "8 traits x 3 marker-count settings (2K, 10K, 50K)"),
        ("scope", "Result6_tabpfn", "10K SNP + 8 traits"),
        ("sheet", "RAW_main36", "Main non-pig manuscript result table with derived R1/R2 columns"),
        ("sheet", "RAW_compare36", "Full non-pig compare table including only-prior, dual, triple, and R2"),
        ("sheet", "R1_summary / R1_cases", "Support tables for Result 1"),
        ("sheet", "R2_*", "Support tables for Result 2"),
        ("sheet", "R3_*", "Support tables for Result 3"),
        ("sheet", "R4_*", "Support tables for sample-size section"),
        ("sheet", "R5_*", "Support tables for marker-count section"),
        ("sheet", "R6_*", "Support tables for TabPFN extensibility section"),
    ]
    return pd.DataFrame(rows, columns=["section", "item", "value"])


def build_r1_tables(main: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    r1_main = main[
        [
            "dataset",
            "trait_slug",
            "no_prior_tabicl",
            "BayesB",
            "GBLUP",
            "RKHS",
        ]
    ].copy()
    for base in ["BayesB", "GBLUP", "RKHS"]:
        r1_main[f"no_vs_{base}_pct"] = pct_diff(r1_main["no_prior_tabicl"], r1_main[base])
        r1_main[f"no_gt_{base}"] = r1_main["no_prior_tabicl"] > r1_main[base]

    r1_summary_rows = []
    for base in ["BayesB", "GBLUP", "RKHS"]:
        r1_summary_rows.append(
            {
                "baseline": base,
                "n_traits": len(main),
                "mean_no_prior_tabicl": main["no_prior_tabicl"].mean(),
                "mean_baseline": main[base].mean(),
                "mean_relative_diff_pct": pct_diff(main["no_prior_tabicl"], main[base]).mean(),
                "no_prior_win_count": int((main["no_prior_tabicl"] > main[base]).sum()),
            }
        )
    r1_summary = pd.DataFrame(r1_summary_rows)

    case_frames = []
    for dataset, trait_slug in R1_CASE_TRAITS:
        row = r1_main[(r1_main["dataset"] == dataset) & (r1_main["trait_slug"] == trait_slug)].copy()
        case_frames.append(row)
    r1_cases = pd.concat(case_frames, ignore_index=True)
    return r1_main, r1_summary, r1_cases


def build_r2_single_long(main: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    prior_map = {
        "BayesB": "single_bayesb_two_step_ls",
        "GBLUP": "single_gblup_two_step_ls",
        "RKHS": "single_rkhs_two_step_ls",
    }
    rows = []
    for prior_name, single_name in prior_map.items():
        block = main[
            [
                "dataset",
                "trait_slug",
                "best_baseline_name",
                "best_baseline",
                "no_prior_tabicl",
                prior_name,
                single_name,
            ]
        ].copy()
        block["prior_name"] = prior_name
        block["single_name"] = single_name
        block["prior_only"] = block[prior_name]
        block["single_value"] = block[single_name]
        block["single_vs_own_prior_abs"] = block["single_value"] - block["prior_only"]
        block["single_vs_own_prior_pct"] = pct_diff(block["single_value"], block["prior_only"])
        block["single_vs_best_baseline_pct"] = pct_diff(block["single_value"], block["best_baseline"])
        block["single_vs_no_prior_pct"] = pct_diff(block["single_value"], block["no_prior_tabicl"])
        rows.append(
            block[
                [
                    "dataset",
                    "trait_slug",
                    "prior_name",
                    "single_name",
                    "best_baseline_name",
                    "no_prior_tabicl",
                    "prior_only",
                    "single_value",
                    "single_vs_own_prior_abs",
                    "single_vs_own_prior_pct",
                    "single_vs_best_baseline_pct",
                    "single_vs_no_prior_pct",
                ]
            ]
        )
    single_long = pd.concat(rows, ignore_index=True)

    summary = (
        single_long.groupby(["prior_name", "single_name"], as_index=False)
        .agg(
            n_traits=("trait_slug", "count"),
            mean_single_vs_own_prior_pct=("single_vs_own_prior_pct", "mean"),
            win_count_vs_own_prior=("single_vs_own_prior_pct", lambda x: int((x > 0).sum())),
        )
        .sort_values("prior_name")
    )

    case_mask = pd.Series(False, index=single_long.index)
    for dataset, trait_slug in R2_SINGLE_CASE_TRAITS:
        case_mask |= (single_long["dataset"] == dataset) & (single_long["trait_slug"] == trait_slug)
    cases = single_long[case_mask].copy()
    return single_long, summary, cases


def build_r2_triple_tables(main: pd.DataFrame, comp: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    triple_trait = main[
        [
            "dataset",
            "trait_slug",
            "best_baseline_name",
            "best_baseline",
            "no_prior_tabicl",
            "BayesB",
            "GBLUP",
            "RKHS",
            "single_bayesb_two_step_ls",
            "single_gblup_two_step_ls",
            "single_rkhs_two_step_ls",
            "triple_two_step_ls",
        ]
    ].copy()
    comp_merge = comp[["dataset", "trait_slug", "only_triple"]].copy()
    triple_trait = triple_trait.merge(comp_merge, on=["dataset", "trait_slug"], how="left")
    for ref in ["BayesB", "GBLUP", "RKHS", "best_baseline", "no_prior_tabicl", "single_bayesb_two_step_ls", "single_gblup_two_step_ls", "single_rkhs_two_step_ls"]:
        triple_trait[f"triple_vs_{ref}_pct"] = pct_diff(triple_trait["triple_two_step_ls"], triple_trait[ref])
    triple_trait["triple_vs_only_triple_pct"] = pct_diff(triple_trait["triple_two_step_ls"], triple_trait["only_triple"])

    summary_rows = []
    for base in ["BayesB", "GBLUP", "RKHS"]:
        diff = pct_diff(main["triple_two_step_ls"], main[base])
        summary_rows.append(
            {
                "block": "triple_vs_baseline",
                "comparison": base,
                "mean_pct": diff.mean(),
                "strict_win_count": int((main["triple_two_step_ls"] > main[base]).sum()),
                "nonloss_count": int((main["triple_two_step_ls"] >= main[base]).sum()),
                "n_traits": len(main),
            }
        )
    for single in ["single_bayesb_two_step_ls", "single_gblup_two_step_ls", "single_rkhs_two_step_ls"]:
        diff = pct_diff(main["triple_two_step_ls"], main[single])
        summary_rows.append(
            {
                "block": "triple_vs_single",
                "comparison": single,
                "mean_pct": diff.mean(),
                "strict_win_count": int((main["triple_two_step_ls"] > main[single]).sum()),
                "nonloss_count": int((main["triple_two_step_ls"] >= main[single]).sum()),
                "n_traits": len(main),
            }
        )

    summary_rows.extend(
        [
            {
                "block": "triple_vs_only_triple",
                "comparison": "Pearson",
                "mean_pct": pct_diff(comp["triple_two_step_ls"], comp["only_triple"]).mean(),
                "strict_win_count": int((comp["triple_two_step_ls"] > comp["only_triple"]).sum()),
                "nonloss_count": int((comp["triple_two_step_ls"] >= comp["only_triple"]).sum()),
                "n_traits": len(comp),
            },
            {
                "block": "triple_vs_only_triple",
                "comparison": "R2",
                "mean_pct": pct_diff(comp["triple_two_step_ls_r2"], comp["only_triple_r2"]).mean(),
                "strict_win_count": int((comp["triple_two_step_ls_r2"] > comp["only_triple_r2"]).sum()),
                "nonloss_count": int((comp["triple_two_step_ls_r2"] >= comp["only_triple_r2"]).sum()),
                "n_traits": len(comp),
            },
            {
                "block": "triple_vs_best",
                "comparison": "best_baseline",
                "mean_pct": pct_diff(main["triple_two_step_ls"], main["best_baseline"]).mean(),
                "strict_win_count": int((main["triple_two_step_ls"] > main["best_baseline"]).sum()),
                "nonloss_count": int((main["triple_two_step_ls"] >= main["best_baseline"]).sum()),
                "n_traits": len(main),
            },
            {
                "block": "triple_vs_no_prior",
                "comparison": "no_prior_tabicl",
                "mean_pct": pct_diff(main["triple_two_step_ls"], main["no_prior_tabicl"]).mean(),
                "strict_win_count": int((main["triple_two_step_ls"] > main["no_prior_tabicl"]).sum()),
                "nonloss_count": int((main["triple_two_step_ls"] >= main["no_prior_tabicl"]).sum()),
                "n_traits": len(main),
            },
        ]
    )
    compare_summary = pd.DataFrame(summary_rows)

    dataset_summary = (
        main.groupby("dataset", as_index=False)
        .apply(
            lambda g: pd.Series(
                {
                    "n_traits": len(g),
                    "triple_vs_best_baseline_pct_mean": pct_diff(g["triple_two_step_ls"], g["best_baseline"]).mean(),
                    "triple_win_count_vs_best": int((g["triple_two_step_ls"] > g["best_baseline"]).sum()),
                    "no_prior_vs_best_baseline_pct_mean": pct_diff(g["no_prior_tabicl"], g["best_baseline"]).mean(),
                    "triple_vs_no_prior_pct_mean": pct_diff(g["triple_two_step_ls"], g["no_prior_tabicl"]).mean(),
                }
            ),
            include_groups=False,
        )
        .reset_index()
        .drop(columns=["index"])
    )

    besttype_summary = (
        main.groupby("best_baseline_name", as_index=False)
        .apply(
            lambda g: pd.Series(
                {
                    "n_traits": len(g),
                    "triple_vs_best_baseline_pct_mean": pct_diff(g["triple_two_step_ls"], g["best_baseline"]).mean(),
                    "triple_win_count_vs_best": int((g["triple_two_step_ls"] > g["best_baseline"]).sum()),
                    "triple_loss_count_vs_best": int((g["triple_two_step_ls"] <= g["best_baseline"]).sum()),
                }
            ),
            include_groups=False,
        )
        .reset_index()
        .drop(columns=["index"])
    )

    triple_vs_best = main[
        ["dataset", "trait_slug", "best_baseline_name", "best_baseline", "triple_two_step_ls"]
    ].copy()
    triple_vs_best["triple_minus_best_abs"] = triple_vs_best["triple_two_step_ls"] - triple_vs_best["best_baseline"]
    triple_vs_best["triple_minus_best_pct"] = pct_diff(
        triple_vs_best["triple_two_step_ls"], triple_vs_best["best_baseline"]
    )
    triple_vs_best["triple_better"] = triple_vs_best["triple_minus_best_pct"] >= 0
    return triple_trait, compare_summary, dataset_summary, besttype_summary, triple_vs_best


def build_r3_tables(comp: pd.DataFrame, ana: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    rep_traits = comp[comp["trait_slug"].isin(["heading_date", "plant_height"])].copy()
    rep_traits["delta_single_bayesb_abs"] = rep_traits["single_bayesb_two_step_ls"] - rep_traits["only_single_bayesb"]
    rep_traits["delta_single_gblup_abs"] = rep_traits["single_gblup_two_step_ls"] - rep_traits["only_single_gblup"]
    rep_traits["delta_single_rkhs_abs"] = rep_traits["single_rkhs_two_step_ls"] - rep_traits["only_single_rkhs"]
    rep_traits["delta_dual_abs"] = rep_traits["dual_two_step_ls"] - rep_traits["only_dual"]
    rep_traits["delta_triple_abs"] = rep_traits["triple_two_step_ls"] - rep_traits["only_triple"]
    rep_traits = rep_traits.merge(
        ana[
            [
                "dataset",
                "trait_slug",
                "w_single_bayesb",
                "w_single_gblup",
                "w_single_rkhs",
                "w_triple",
                "bayesb_share",
                "gblup_share",
                "rkhs_share",
            ]
        ],
        on=["dataset", "trait_slug"],
        how="left",
    )

    weights36 = ana[
        [
            "dataset",
            "trait_slug",
            "best_baseline_name",
            "no_prior_tabicl",
            "BayesB",
            "GBLUP",
            "RKHS",
            "single_bayesb_two_step_ls",
            "single_gblup_two_step_ls",
            "single_rkhs_two_step_ls",
            "triple_two_step_ls",
            "w_single_bayesb",
            "w_single_gblup",
            "w_single_rkhs",
            "w_triple",
            "bayesb_share",
            "gblup_share",
            "rkhs_share",
            "baseline_entropy",
            "baseline_spread",
        ]
    ].copy()
    best_vals = weights36[["BayesB", "GBLUP", "RKHS"]].max(axis=1)
    weights36["no_prior_vs_best_baseline_pct"] = pct_diff(weights36["no_prior_tabicl"], best_vals)
    weights36["triple_vs_best_baseline_pct"] = pct_diff(weights36["triple_two_step_ls"], best_vals)

    no_vs_best = pct_diff(ana["no_prior_tabicl"], ana[["BayesB", "GBLUP", "RKHS"]].max(axis=1))
    triple_vs_best = pct_diff(ana["triple_two_step_ls"], ana[["BayesB", "GBLUP", "RKHS"]].max(axis=1))
    q = pd.qcut(ana["w_triple"], 3, labels=["low", "mid", "high"])

    weight_summary_rows = []
    for col in ["w_single_bayesb", "w_single_gblup", "w_single_rkhs", "w_triple"]:
        weight_summary_rows.append({"metric": f"{col}_mean", "value": ana[col].mean()})
        weight_summary_rows.append({"metric": f"{col}_min", "value": ana[col].min()})
        weight_summary_rows.append({"metric": f"{col}_max", "value": ana[col].max()})
    weight_summary_rows.append({"metric": "corr_w_triple_vs_no_prior_vs_best_pct", "value": ana["w_triple"].corr(no_vs_best)})
    for label in ["low", "mid", "high"]:
        idx = q == label
        weight_summary_rows.append({"metric": f"{label}_tertile_triple_vs_best_pct_mean", "value": triple_vs_best[idx].mean()})
        weight_summary_rows.append({"metric": f"{label}_tertile_no_prior_vs_best_pct_mean", "value": no_vs_best[idx].mean()})
    weight_summary = pd.DataFrame(weight_summary_rows)

    quadrant_map = {
        "single_bayesb_two_step_ls": "only_single_bayesb",
        "single_gblup_two_step_ls": "only_single_gblup",
        "single_rkhs_two_step_ls": "only_single_rkhs",
        "dual_two_step_ls": "only_dual",
        "triple_two_step_ls": "only_triple",
    }
    quadrant_rows = []
    for fusion_col, prior_only_col in quadrant_map.items():
        x = comp[fusion_col] - comp["no_prior_tabicl"]
        y = comp[fusion_col] - comp[prior_only_col]
        quadrant_rows.append(
            {
                "fusion_model": fusion_col,
                "prior_only_model": prior_only_col,
                "double_positive_count": int(((x > 0) & (y > 0)).sum()),
                "only_vs_no_prior_positive_count": int(((x > 0) & (y <= 0)).sum()),
                "only_vs_prior_positive_count": int(((x <= 0) & (y > 0)).sum()),
                "double_negative_or_tie_count": int(((x <= 0) & (y <= 0)).sum()),
                "n_traits": len(comp),
            }
        )
    quadrants = pd.DataFrame(quadrant_rows)

    mean3 = ana[["BayesB", "GBLUP", "RKHS"]].mean(axis=1)
    second_best = ana[["BayesB", "GBLUP", "RKHS"]].apply(lambda row: row.nlargest(2).iloc[-1], axis=1)
    geo_trait = ana[
        [
            "dataset",
            "trait_slug",
            "best_baseline_name",
            "BayesB",
            "GBLUP",
            "RKHS",
            "no_prior_tabicl",
            "triple_two_step_ls",
            "baseline_entropy",
            "baseline_spread",
        ]
    ].copy()
    geo_trait["baseline_mean3"] = mean3
    geo_trait["rel_bayesb"] = (geo_trait["BayesB"] - geo_trait["baseline_mean3"]) / geo_trait["baseline_mean3"]
    geo_trait["rel_gblup"] = (geo_trait["GBLUP"] - geo_trait["baseline_mean3"]) / geo_trait["baseline_mean3"]
    geo_trait["rel_rkhs"] = (geo_trait["RKHS"] - geo_trait["baseline_mean3"]) / geo_trait["baseline_mean3"]
    geo_trait["no_prior_vs_best_pct"] = pct_diff(
        geo_trait["no_prior_tabicl"], geo_trait[["BayesB", "GBLUP", "RKHS"]].max(axis=1)
    )
    geo_trait["no_prior_vs_mean3_pct"] = pct_diff(geo_trait["no_prior_tabicl"], geo_trait["baseline_mean3"])
    geo_trait["triple_vs_best_pct"] = pct_diff(
        geo_trait["triple_two_step_ls"], geo_trait[["BayesB", "GBLUP", "RKHS"]].max(axis=1)
    )
    geo_trait["dominance_margin_pct"] = (geo_trait[["BayesB", "GBLUP", "RKHS"]].max(axis=1) - second_best) / geo_trait["baseline_mean3"] * 100.0

    geo_sum_rows = [
        {"metric": "corr_rel_bayesb_vs_no_prior_vs_best_pct", "value": geo_trait["rel_bayesb"].corr(geo_trait["no_prior_vs_best_pct"])},
        {"metric": "corr_rel_gblup_vs_no_prior_vs_best_pct", "value": geo_trait["rel_gblup"].corr(geo_trait["no_prior_vs_best_pct"])},
        {"metric": "corr_rel_rkhs_vs_no_prior_vs_best_pct", "value": geo_trait["rel_rkhs"].corr(geo_trait["no_prior_vs_best_pct"])},
        {"metric": "corr_dominance_margin_vs_no_prior_vs_mean3_pct", "value": geo_trait["dominance_margin_pct"].corr(geo_trait["no_prior_vs_mean3_pct"])},
        {"metric": "corr_entropy_vs_no_prior_vs_best_pct", "value": geo_trait["baseline_entropy"].corr(geo_trait["no_prior_vs_best_pct"])},
        {"metric": "corr_entropy_vs_triple_vs_best_pct", "value": geo_trait["baseline_entropy"].corr(geo_trait["triple_vs_best_pct"])},
        {"metric": "corr_spread_vs_no_prior_vs_best_pct", "value": geo_trait["baseline_spread"].corr(geo_trait["no_prior_vs_best_pct"])},
        {"metric": "corr_spread_vs_triple_vs_best_pct", "value": geo_trait["baseline_spread"].corr(geo_trait["triple_vs_best_pct"])},
    ]

    by_besttype = (
        geo_trait.groupby("best_baseline_name", as_index=False)
        .apply(
            lambda g: pd.Series(
                {
                    "n_traits": len(g),
                    "no_prior_vs_best_pct_mean": g["no_prior_vs_best_pct"].mean(),
                    "no_prior_vs_mean3_pct_mean": g["no_prior_vs_mean3_pct"].mean(),
                    "no_prior_win_count_vs_best": int((g["no_prior_tabicl"] > g[["BayesB", "GBLUP", "RKHS"]].max(axis=1)).sum()),
                }
            ),
            include_groups=False,
        )
        .reset_index()
        .drop(columns=["index"])
    )
    geo_sum = pd.concat([pd.DataFrame(geo_sum_rows), by_besttype], ignore_index=False)
    return rep_traits, weights36, weight_summary, quadrants, geo_trait, geo_sum


def build_sample_size_tables(main: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    raw = load_sample_size_raw()
    raw = raw[raw[["dataset", "trait_slug"]].apply(tuple, axis=1).isin(SAMPLE_SIZE_TRAITS_8)].copy()
    keep_cols = [
        "dataset",
        "trait_slug",
        "sample_fraction",
        "sample_label",
        "repeat",
        "no_prior_tabicl",
        "BayesB",
        "GBLUP",
        "RKHS",
        "single_bayesb_two_step_ls",
        "single_gblup_two_step_ls",
        "single_rkhs_two_step_ls",
        "triple_two_step_ls",
    ]
    raw = raw[keep_cols]
    numeric_cols = [c for c in raw.columns if c not in {"dataset", "trait_slug", "sample_fraction", "sample_label", "repeat"}]
    sample_trait = (
        raw.groupby(["dataset", "trait_slug", "sample_fraction", "sample_label"], as_index=False)[numeric_cols]
        .mean()
        .sort_values(["dataset", "trait_slug", "sample_fraction"])
    )
    repeat_count = raw.groupby(["dataset", "trait_slug", "sample_fraction"], as_index=False).agg(n_repeats=("repeat", "count"))
    sample_trait = sample_trait.merge(repeat_count, on=["dataset", "trait_slug", "sample_fraction"], how="left")
    sample_trait["source"] = "sample_size_server_dir_repeat_mean"

    main_100 = main[main[["dataset", "trait_slug"]].apply(tuple, axis=1).isin(SAMPLE_SIZE_TRAITS_8)].copy()
    main_100 = main_100[
        [
            "dataset",
            "trait_slug",
            "no_prior_tabicl",
            "BayesB",
            "GBLUP",
            "RKHS",
            "single_bayesb_two_step_ls",
            "single_gblup_two_step_ls",
            "single_rkhs_two_step_ls",
            "triple_two_step_ls",
        ]
    ]
    main_100["sample_fraction"] = "1.00"
    main_100["sample_label"] = "100%"
    main_100["n_repeats"] = 1
    main_100["source"] = "main_results_non_pig_fixed"

    sample_trait = pd.concat([sample_trait, main_100], ignore_index=True, sort=False)
    sample_trait["sample_sort"] = sample_trait["sample_fraction"].astype(float)
    sample_trait = sample_trait.sort_values(["dataset", "trait_slug", "sample_sort"]).drop(columns=["sample_sort"])

    best_name, best_value = best_of_three(sample_trait, ["BayesB", "GBLUP", "RKHS"])
    sample_trait["best_baseline_name"] = best_name
    sample_trait["best_baseline"] = best_value
    for base in ["BayesB", "GBLUP", "RKHS", "best_baseline"]:
        sample_trait[f"triple_vs_{base}_pct"] = pct_diff(sample_trait["triple_two_step_ls"], sample_trait[base])

    sample_mean = (
        sample_trait.groupby(["sample_fraction", "sample_label"], as_index=False)
        .agg(
            n_traits=("trait_slug", "count"),
            mean_no_prior=("no_prior_tabicl", "mean"),
            mean_BayesB=("BayesB", "mean"),
            mean_GBLUP=("GBLUP", "mean"),
            mean_RKHS=("RKHS", "mean"),
            mean_single_bayesb=("single_bayesb_two_step_ls", "mean"),
            mean_single_gblup=("single_gblup_two_step_ls", "mean"),
            mean_single_rkhs=("single_rkhs_two_step_ls", "mean"),
            mean_triple=("triple_two_step_ls", "mean"),
            mean_triple_vs_BayesB_pct=("triple_vs_BayesB_pct", "mean"),
            mean_triple_vs_GBLUP_pct=("triple_vs_GBLUP_pct", "mean"),
            mean_triple_vs_RKHS_pct=("triple_vs_RKHS_pct", "mean"),
            mean_triple_vs_best_pct=("triple_vs_best_baseline_pct", "mean"),
            triple_gt_best_count=("triple_vs_best_baseline_pct", lambda x: int((x > 0).sum())),
        )
        .sort_values("sample_fraction")
    )

    case_traits = {
        ("rice529", "grain_weight"),
        ("rice529", "grain_width"),
        ("soybean951", "lw_beijing_2013_lw_beijing_2013"),
        ("wheat406", "sl_e2"),
    }
    sample_cases = sample_trait[
        sample_trait[["dataset", "trait_slug"]].apply(tuple, axis=1).isin(case_traits)
    ].copy()
    sample_cases = sample_cases[
        [
            "dataset",
            "trait_slug",
            "sample_fraction",
            "sample_label",
            "no_prior_tabicl",
            "BayesB",
            "GBLUP",
            "RKHS",
            "single_bayesb_two_step_ls",
            "single_gblup_two_step_ls",
            "single_rkhs_two_step_ls",
            "triple_two_step_ls",
            "best_baseline_name",
            "best_baseline",
            "triple_vs_BayesB_pct",
            "triple_vs_GBLUP_pct",
            "triple_vs_RKHS_pct",
            "triple_vs_best_baseline_pct",
        ]
    ].sort_values(["dataset", "trait_slug", "sample_fraction"])
    return sample_trait, sample_mean, sample_cases


def build_marker_tables(marker: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    marker = marker.copy()
    marker["marker_count"] = marker["marker_count"].astype(int)
    marker["marker_label"] = marker["marker_count"].map({2000: "2K", 10000: "10K", 50000: "50K"})

    marker_mean = (
        marker.groupby(["marker_count", "marker_label"], as_index=False)
        .agg(
            n_traits=("trait_slug", "count"),
            mean_no_prior=("no_prior_tabicl", "mean"),
            mean_BayesB=("BayesB", "mean"),
            mean_GBLUP=("GBLUP", "mean"),
            mean_RKHS=("RKHS", "mean"),
            mean_single_bayesb=("single_bayesb_two_step_ls", "mean"),
            mean_single_gblup=("single_gblup_two_step_ls", "mean"),
            mean_single_rkhs=("single_rkhs_two_step_ls", "mean"),
            mean_triple=("triple_two_step_ls", "mean"),
            mean_triple_vs_BayesB_pct=("triple_two_step_ls_vs_BayesB_pct", "mean"),
            mean_triple_vs_GBLUP_pct=("triple_two_step_ls_vs_GBLUP_pct", "mean"),
            mean_triple_vs_RKHS_pct=("triple_two_step_ls_vs_RKHS_pct", "mean"),
            mean_single_bayesb_vs_own_prior_pct=("single_bayesb_two_step_ls_vs_own_prior_pct", "mean"),
            mean_single_gblup_vs_own_prior_pct=("single_gblup_two_step_ls_vs_own_prior_pct", "mean"),
            mean_single_rkhs_vs_own_prior_pct=("single_rkhs_two_step_ls_vs_own_prior_pct", "mean"),
            triple_gt_best_count=("triple_two_step_ls_vs_best_baseline_pct", lambda x: int((x > 0).sum())),
        )
        .sort_values("marker_count")
    )

    triple_trait = marker[
        [
            "dataset",
            "trait_slug",
            "marker_count",
            "marker_label",
            "best_baseline_name",
            "best_baseline",
            "no_prior_tabicl",
            "BayesB",
            "GBLUP",
            "RKHS",
            "single_bayesb_two_step_ls",
            "single_gblup_two_step_ls",
            "single_rkhs_two_step_ls",
            "triple_two_step_ls",
            "triple_two_step_ls_vs_best_baseline_pct",
            "triple_two_step_ls_vs_BayesB_pct",
            "triple_two_step_ls_vs_GBLUP_pct",
            "triple_two_step_ls_vs_RKHS_pct",
        ]
    ].copy()
    triple_trend = triple_trait.pivot_table(
        index=["dataset", "trait_slug"],
        columns="marker_label",
        values="triple_two_step_ls",
        aggfunc="first",
    ).reset_index()
    triple_trend["nondecreasing_2k_10k_50k"] = (triple_trend["2K"] <= triple_trend["10K"]) & (triple_trend["10K"] <= triple_trend["50K"])
    marker_cases = triple_trait[
        triple_trait[["dataset", "trait_slug"]].apply(tuple, axis=1).isin(
            {
                ("cotton1245", "cotton_fibelo_17_18_cotton_fibelo_17_18"),
                ("rice529", "grain_weight"),
                ("rice529", "grain_width"),
                ("soybean951", "lw_beijing_2013_lw_beijing_2013"),
                ("wheat406", "sl_e2"),
            }
        )
    ].copy()
    marker_cases = marker_cases.sort_values(["dataset", "trait_slug", "marker_count"])
    return marker, marker_mean, triple_trend, marker_cases


def build_tabpfn_tables(tabpfn: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    tabpfn = tabpfn.copy()
    pos_traits = tabpfn.loc[tabpfn["triple_vs_best_baseline_pct"] > 0, "trait"].tolist()
    neg_traits = tabpfn.loc[tabpfn["triple_vs_best_baseline_pct"] <= 0, "trait"].tolist()

    summary_rows = [
        {"metric": "mean_no_prior_tabpfn", "value": tabpfn["no_prior_tabpfn"].mean()},
        {"metric": "mean_BayesB", "value": tabpfn["BayesB"].mean()},
        {"metric": "mean_GBLUP", "value": tabpfn["GBLUP"].mean()},
        {"metric": "mean_RKHS", "value": tabpfn["RKHS"].mean()},
        {"metric": "mean_only_triple_fusion", "value": tabpfn["only_triple_fusion"].mean()},
        {"metric": "mean_triple_fusion", "value": tabpfn["triple_fusion"].mean()},
        {"metric": "mean_only_vs_BayesB_pct", "value": tabpfn["only_vs_BayesB_pct"].mean()},
        {"metric": "mean_only_vs_GBLUP_pct", "value": tabpfn["only_vs_GBLUP_pct"].mean()},
        {"metric": "mean_only_vs_RKHS_pct", "value": tabpfn["only_vs_RKHS_pct"].mean()},
        {"metric": "mean_triple_vs_BayesB_pct", "value": tabpfn["triple_vs_BayesB_pct"].mean()},
        {"metric": "mean_triple_vs_GBLUP_pct", "value": tabpfn["triple_vs_GBLUP_pct"].mean()},
        {"metric": "mean_triple_vs_RKHS_pct", "value": tabpfn["triple_vs_RKHS_pct"].mean()},
        {"metric": "mean_triple_vs_best_baseline_pct", "value": tabpfn["triple_vs_best_baseline_pct"].mean()},
        {"metric": "mean_triple_vs_only_triple_pct", "value": tabpfn["triple_vs_only_triple_pct"].mean()},
        {"metric": "triple_positive_trait_count_vs_best", "value": int((tabpfn["triple_vs_best_baseline_pct"] > 0).sum())},
        {"metric": "triple_negative_trait_count_vs_best", "value": int((tabpfn["triple_vs_best_baseline_pct"] <= 0).sum())},
        {"metric": "positive_traits_vs_best", "value": "; ".join(pos_traits)},
        {"metric": "negative_traits_vs_best", "value": "; ".join(neg_traits)},
    ]
    return tabpfn, pd.DataFrame(summary_rows)


def build_tabpfn_extended_tables(tabpfn: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    prior_dir_map = {
        "BayesB": ("prior_only_bayesb", "tabicl_bayesb_single_prior"),
        "GBLUP": ("prior_only_gblup", "tabicl_gblup_single_prior"),
        "RKHS": ("prior_only_rkhs", "tabicl_rkhs_single_prior"),
    }
    rows = []
    weight_rows = []
    for dataset_dir in sorted(TABPFN_SUPPORT_DIR.iterdir()):
        if not dataset_dir.is_dir():
            continue
        dataset = dataset_dir.name
        for trait_dir in sorted(dataset_dir.iterdir()):
            if not trait_dir.is_dir():
                continue
            trait = trait_dir.name
            sum_path = trait_dir / "decoupled_fusion_summary.json"
            if not sum_path.exists():
                continue
            meta = json.loads(sum_path.read_text())
            for prior_name, (prior_only_dir, fusion_dir) in prior_dir_map.items():
                prior_only_summary = json.loads((trait_dir / prior_only_dir / "summary.json").read_text())
                fusion_summary = json.loads((trait_dir / fusion_dir / "summary.json").read_text())
                rows.append(
                    {
                        "dataset": dataset,
                        "trait": trait,
                        "prior_name": prior_name,
                        "prior_only_value": prior_only_summary["pearson_mean"],
                        "single_fusion_value": fusion_summary["pearson_mean"],
                        "single_vs_own_prior_pct": (fusion_summary["pearson_mean"] - prior_only_summary["pearson_mean"])
                        / prior_only_summary["pearson_mean"]
                        * 100.0,
                        "w_tabpfn": meta["single"][prior_name]["w_tabicl"],
                    }
                )
                weight_rows.append(
                    {
                        "dataset": dataset,
                        "trait": trait,
                        "fusion_type": f"single_{prior_name.lower()}",
                        "w_tabpfn": meta["single"][prior_name]["w_tabicl"],
                        "bayesb_share": 1.0 if prior_name == "BayesB" else 0.0,
                        "gblup_share": 1.0 if prior_name == "GBLUP" else 0.0,
                        "rkhs_share": 1.0 if prior_name == "RKHS" else 0.0,
                    }
                )
            weight_rows.append(
                {
                    "dataset": dataset,
                    "trait": trait,
                    "fusion_type": "triple",
                    "w_tabpfn": meta["triple"]["w_tabicl"],
                    "bayesb_share": meta["triple"]["prior_weights"]["BayesB"],
                    "gblup_share": meta["triple"]["prior_weights"]["GBLUP"],
                    "rkhs_share": meta["triple"]["prior_weights"]["RKHS"],
                }
            )

    single_long = pd.DataFrame(rows)
    weight_df = pd.DataFrame(weight_rows)
    single_summary = (
        single_long.groupby("prior_name", as_index=False)
        .agg(
            n_traits=("trait", "count"),
            mean_prior_only=("prior_only_value", "mean"),
            mean_single_fusion=("single_fusion_value", "mean"),
            mean_single_vs_own_prior_pct=("single_vs_own_prior_pct", "mean"),
            mean_w_tabpfn=("w_tabpfn", "mean"),
            single_win_count=("single_vs_own_prior_pct", lambda x: int((x > 0).sum())),
        )
        .sort_values("prior_name")
    )
    return single_long, single_summary, weight_df


def autofit_and_style(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    top = Alignment(vertical="top")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row >= 1:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = top
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            values = ["" if cell.value is None else str(cell.value) for cell in col_cells[:200]]
            max_len = max((len(v) for v in values), default=0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
    wb.save(path)


def build_workbook() -> Path:
    for path in [MAIN_RESULTS_PATH, COMPARE_ALL_PATH, MARKER_COUNT_PATH, TABPFN_PATH, ANALYSIS_TABLE_PATH]:
        require_path(path)
    require_path(SAMPLE_SIZE_DIR)
    require_path(TABPFN_SUPPORT_DIR)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    main = pd.read_csv(MAIN_RESULTS_PATH)
    comp_all = pd.read_csv(COMPARE_ALL_PATH)
    comp = comp_all[comp_all["dataset"] != "pig3534"].copy()
    marker = pd.read_csv(MARKER_COUNT_PATH)
    tabpfn = pd.read_csv(TABPFN_PATH)
    ana = pd.read_csv(ANALYSIS_TABLE_PATH)

    raw_main36 = main.copy()
    raw_compare36 = comp.copy()

    r1_main, r1_summary, r1_cases = build_r1_tables(main)
    r2_single_long, r2_single_summary, r2_single_cases = build_r2_single_long(main)
    r2_triple_trait, r2_compare_sum, r2_dataset, r2_besttype, r2_strongest_single = build_r2_triple_tables(main, comp)
    r3_rep_traits, r3_weights36, r3_weight_sum, r3_quadrants, r3_geo_trait, r3_geo_sum = build_r3_tables(comp, ana)
    r4_sample_trait, r4_sample_mean, r4_sample_cases = build_sample_size_tables(main)
    r5_marker_raw, r5_marker_mean, r5_marker_trait, r5_marker_cases = build_marker_tables(marker)
    r6_tabpfn_raw, r6_tabpfn_sum = build_tabpfn_tables(tabpfn)
    r6_tabpfn_single_long, r6_tabpfn_single_sum, r6_tabpfn_weights = build_tabpfn_extended_tables(tabpfn)

    sheets = [
        ("README", round_frame(build_readme_sheet(), 6)),
        ("RAW_main36", round_frame(raw_main36, 10)),
        ("RAW_compare36", round_frame(raw_compare36, 10)),
        ("R1_main36", round_frame(r1_main, 6)),
        ("R1_summary", round_frame(r1_summary, 6)),
        ("R1_cases", round_frame(r1_cases, 6)),
        ("R2_single_long", round_frame(r2_single_long, 6)),
        ("R2_single_sum", round_frame(r2_single_summary, 6)),
        ("R2_single_cases", round_frame(r2_single_cases, 6)),
        ("R2_triple_trait", round_frame(r2_triple_trait, 6)),
        ("R2_compare_sum", round_frame(r2_compare_sum, 6)),
        ("R2_dataset", round_frame(r2_dataset, 6)),
        ("R2_besttype", round_frame(r2_besttype, 6)),
        ("R2_strongest_single", round_frame(r2_strongest_single, 6)),
        ("R3_rep_traits", round_frame(r3_rep_traits, 6)),
        ("R3_weights36", round_frame(r3_weights36, 6)),
        ("R3_weight_sum", round_frame(r3_weight_sum, 6)),
        ("R3_quadrants", round_frame(r3_quadrants, 6)),
        ("R3_geo_trait", round_frame(r3_geo_trait, 6)),
        ("R3_geo_sum", round_frame(r3_geo_sum, 6)),
        ("R4_sample8_trait", round_frame(r4_sample_trait, 6)),
        ("R4_sample8_mean", round_frame(r4_sample_mean, 6)),
        ("R4_sample8_cases", round_frame(r4_sample_cases, 6)),
        ("R5_marker_raw", round_frame(r5_marker_raw, 10)),
        ("R5_marker_mean", round_frame(r5_marker_mean, 6)),
        ("R5_marker_trait", round_frame(r5_marker_trait, 6)),
        ("R5_marker_cases", round_frame(r5_marker_cases, 6)),
        ("R6_tabpfn_raw", round_frame(r6_tabpfn_raw, 10)),
        ("R6_tabpfn_sum", round_frame(r6_tabpfn_sum, 6)),
        ("R6_tabpfn_single", round_frame(r6_tabpfn_single_long, 6)),
        ("R6_tabpfn_single_sum", round_frame(r6_tabpfn_single_sum, 6)),
        ("R6_tabpfn_weights", round_frame(r6_tabpfn_weights, 6)),
    ]

    with pd.ExcelWriter(WORKBOOK_PATH, engine="openpyxl") as writer:
        for sheet_name, df in sheets:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    autofit_and_style(WORKBOOK_PATH)
    return WORKBOOK_PATH


if __name__ == "__main__":
    output_path = build_workbook()
    print(output_path)
